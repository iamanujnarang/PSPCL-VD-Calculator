import streamlit as st
import pandas as pd
import numpy as np
import io
import graphviz
from fpdf import FPDF
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------- PAGE CONFIG ----------------
st.set_page_config(
    page_title="PSPCL VD Calculator Pro",
    page_icon="⚡",
    layout="wide"
)

# ---------------- CONSTANTS ----------------
# Updated Logo URL (Direct PNG Link)
PSPCL_LOGO = "https://pspcl.in/assets/images/logo.png"
INSTA_ICON = "https://upload.wikimedia.org/wikipedia/commons/a/a5/Instagram_icon.png"
FB_ICON = "https://upload.wikimedia.org/wikipedia/commons/1/1b/Facebook_icon.svg"
X_LOGO_FINAL = "https://upload.wikimedia.org/wikipedia/commons/b/b7/X_logo.jpg"
LINKEDIN_ICON = "https://upload.wikimedia.org/wikipedia/commons/c/ca/LinkedIn_logo_initials.png"
BEECLUE_LOGO = "https://raw.githubusercontent.com/iamanujnarang/LDHF/e5748e037b76a52a47d610a88c3a3c70f72f1c9a/BEECLUE.png"

# HV LIMITS
HV_UPPER_LIMIT = 6.0
HV_LOWER_LIMIT = -9.0

VD_FACTORS = {
    "ACSR 100 SQMM": 0.0415,
    "ACSR 80 SQMM": 0.0512,
    "ACSR 50 SQMM": 0.0910,
    "ACSR 30 SQMM": 0.1520,
    "ACSR 20 SQMM": 0.2250,
    "XLPE CABLE 300 SQMM": 0.0146,
    "XLPE CABLE 150 SQMM": 0.0285,
    "XLPE CABLE 35 SQMM": 0.1150
}

# ---------------- CSS (Updated for Standard Footer) ----------------
st.markdown("""
<style>
    .main { background-color: #f8fafc; }
    .header-box {
        text-align:center;
        padding:25px;
        background:white;
        border-radius:15px;
        margin-bottom:25px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05);
    }
    .pspcl-main-logo { height:120px; }
    
    /* Footer Styling Matching Other Apps */
    .footer-container {
        text-align: center;
        margin-top: 80px;
        padding: 40px 20px;
        border-top: 1px solid #ddd;
    }
    .made-with-love {
        font-size: 1.2rem;
        color: #334155;
        margin-bottom: 20px;
    }
    .heart-symbol { color: #e63946; }
    .social-icon {
        width: 30px;
        margin: 0 10px;
        transition: 0.3s;
    }
    .social-icon:hover { transform: scale(1.2); }
    .powered-text {
        color: #94a3b8;
        font-size: 0.7rem;
        letter-spacing: 2px;
        margin-bottom: 10px;
        text-transform: uppercase;
    }
    .beeclue-img { width: 180px; height: auto; }
</style>
""", unsafe_allow_html=True)

# ---------------- HEADER ----------------
st.markdown(f"""
<div class="header-box">
<img src="{PSPCL_LOGO}" class="pspcl-main-logo">
<h1>PUNJAB STATE POWER CORPORATION LIMITED</h1>
<h3>Voltage Drop Calculator (11kV / 33kV)</h3>
<p>HV Limits: +6% to -9%</p>
</div>
""", unsafe_allow_html=True)

# ---------------- SIDEBAR ----------------
with st.sidebar:
    st.header("⚡ Input Settings")
    feeder = st.text_input("Feeder Name", "FEEDER-01")
    substation = st.text_input("Substation", "MAIN SS")
    mdi_a = st.number_input("MDI (Amps)", value=100.0)
    n_sec = st.number_input("No. of Sections", min_value=1, value=5)
    mdi_kva = round(np.sqrt(3)*11*mdi_a,2)
    st.success(f"MDI = {mdi_kva} kVA")
    st.info("Permissible HV Range: +6% to -9%")

# ---------------- INPUT TABLE ----------------
sections = [f"{chr(65+i)}-{chr(66+i)}" for i in range(int(n_sec))]

df = pd.DataFrame({
    "SECTION": sections,
    "CONDUCTOR SIZE": [None]*int(n_sec),
    "LENGTH (KM)": [0.0]*int(n_sec),
    "NET LOAD (kVA)": [0.0]*int(n_sec)
})

df = st.data_editor(
    df,
    column_config={
        "CONDUCTOR SIZE": st.column_config.SelectboxColumn(options=list(VD_FACTORS.keys()))
    },
    use_container_width=True
)

# ---------------- CALCULATION ----------------
if st.button("🚀 Calculate", use_container_width=True):
    if df["CONDUCTOR SIZE"].isnull().any():
        st.error("Select conductor for all sections")
        st.stop()

    df["FACTOR"] = df["CONDUCTOR SIZE"].map(VD_FACTORS)

    # cumulative load
    loads = df["NET LOAD (kVA)"].tolist()
    cum = []
    running = 0
    for val in loads[::-1]:
        running += val
        cum.append(running)
    df["CUM LOAD"] = cum[::-1]

    df["SECTION VD"] = df["LENGTH (KM)"] * df["CUM LOAD"] * df["FACTOR"]

    sum_vd = df["SECTION VD"].sum()
    max_load = df["CUM LOAD"].iloc[0]

    demand_factor = mdi_kva/max_load if max_load>0 else 0
    actual_vd = sum_vd * demand_factor
    vd_percent = (actual_vd/(11000-actual_vd))*100 if (11000-actual_vd)>0 else 0

    # ---------------- DISPLAY ----------------
    st.subheader("📊 Results")
    st.dataframe(df, use_container_width=True)

    c1, c2, c3 = st.columns(3)
    c1.metric("Demand Factor", f"{demand_factor:.4f}")
    c2.metric("Actual VD (V)", f"{actual_vd:.2f}")
    c3.metric("VD %", f"{vd_percent:.3f}%")

    if vd_percent > HV_UPPER_LIMIT or vd_percent < HV_LOWER_LIMIT:
        st.error(f"⚠️ Voltage {vd_percent:.3f}% OUTSIDE permissible limits (+6% to -9%)")
    else:
        st.success(f"✅ Voltage {vd_percent:.3f}% WITHIN permissible limits (+6% to -9%)")

    # ---------------- GRAPH ----------------
    st.subheader("🔌 Network Diagram")
    dot = graphviz.Digraph()
    dot.node("SOURCE", f"{substation}\n11kV")
    last = "SOURCE"
    for i, row in df.iterrows():
        node = f"N{i}"
        dot.node(node, f"{row['SECTION']}\n{row['NET LOAD (kVA)']} kVA")
        dot.edge(last, node, label=f"{row['LENGTH (KM)']} km")
        last = node
    st.graphviz_chart(dot)

    # ---------------- PDF ----------------
    if st.button("📄 Export PDF"):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(0,10,"PSPCL Voltage Drop Report",ln=True)
        pdf.cell(0,10,f"Feeder: {feeder}",ln=True)
        pdf.cell(0,10,f"VD %: {vd_percent:.3f}",ln=True)
        status = "PASS" if (HV_LOWER_LIMIT <= vd_percent <= HV_UPPER_LIMIT) else "FAIL"
        pdf.cell(0,10,f"Status: {status}",ln=True)
        pdf_bytes = pdf.output(dest='S').encode('latin1')
        st.download_button("Download PDF", pdf_bytes, "VD_Report.pdf")

# ---------------- UPDATED STANDARD FOOTER ----------------
footer_html = f"""
<div class="footer-container">
    <div class="made-with-love">Made with <span class="heart-symbol">❤️</span> by <b>Er. Anuj Narang, JE PSPCL</b></div>
    <div style="margin-bottom: 25px;">
        <a href="https://instagram.com/iamanujnarang" target="_blank"><img src="{INSTA_ICON}" class="social-icon"></a>
        <a href="https://facebook.com/iamanujnarang" target="_blank"><img src="{FB_ICON}" class="social-icon"></a>
        <a href="https://x.com/iamanujnarang" target="_blank"><img src="{X_LOGO_FINAL}" class="social-icon"></a>
        <a href="https://linkedin.com/in/iamanujnarang" target="_blank"><img src="{LINKEDIN_ICON}" class="social-icon"></a>
    </div>
    <div style="margin-top: 25px;">
        <div class="powered-text">In Strategic Collaboration with</div>
        <a href="https://beeclue.com" target="_blank">
            <img src="{BEECLUE_LOGO}" class="beeclue-img">
        </a>
    </div>
    <div style="color: #94a3b8; font-size: 0.85rem; margin-top: 25px;">© 2026 | PSPCL Guidelines | CC 35/2025</div>
</div>
"""
st.markdown(footer_html, unsafe_allow_html=True)
