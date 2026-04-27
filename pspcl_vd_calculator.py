import streamlit as st
import pandas as pd
import numpy as np
import io
import time
import graphviz
from fpdf import FPDF # Backend requires fpdf2
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================
# 1. GLOBAL CONFIGURATION & CONSTANTS
# ==========================================
st.set_page_config(
    page_title="PSPCL 11kV VD Calculator Pro",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Assets (All Links Fixed & Verified)
PSPCL_LOGO = "https://upload.wikimedia.org/wikipedia/en/3/3a/Punjab_State_Power_Corporation_Limited_logo.png"
INSTA_ICON = "https://upload.wikimedia.org/wikipedia/commons/a/a5/Instagram_icon.png"
FB_ICON = "https://upload.wikimedia.org/wikipedia/commons/1/1b/Facebook_icon.svg"
X_LOGO_FINAL = "https://upload.wikimedia.org/wikipedia/commons/b/b7/X_logo.jpg"
LINKEDIN_ICON = "https://upload.wikimedia.org/wikipedia/commons/c/ca/LinkedIn_logo_initials.png"
BEECLUE_LOGO_WHITE = "https://beeclue.com/wp-content/uploads/2026/02/b-horizontal-logo-w-2048x506.png"

VD_FACTORS = {
    "ACSR 100 SQMM (Dog)": 0.0415, "ACSR 80 SQMM (Wolf)": 0.0512, "ACSR 50 SQMM (Rabbit)": 0.0910,
    "ACSR 30 SQMM (Weasel)": 0.1520, "ACSR 20 SQMM (Squirrel)": 0.2250, "XLPE CABLE 300 SQMM": 0.0146,
    "XLPE CABLE 150 SQMM": 0.0285, "XLPE CABLE 35 SQMM": 0.1150
}

# ==========================================
# 2. CUSTOM CSS STYLING
# ==========================================
st.markdown(f"""
<style>
    .main {{ background-color: #f8f9fa; }}
    
    /* Header Section */
    .header-box {{ text-align: center; padding: 30px; background: white; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); margin-bottom: 25px; border-top: 8px solid #ffcc00; }}
    .pspcl-main-logo {{ height: 130px; margin-bottom: 15px; }}
    
    /* Metrics Box */
    .formula-section {{ background: #ffffff; padding: 30px; border-radius: 15px; border: 1px solid #dee2e6; margin-top: 30px; box-shadow: 0 10px 25px rgba(0,0,0,0.05); }}
    
    /* Footer Styling with Dark Background for Beeclue */
    .footer-container {{ text-align: center; margin-top: 60px; padding: 60px 20px; border-top: 1px solid #eee; background: #ffffff; border-radius: 20px 20px 0 0;}}
    .social-logo {{ width: 42px; margin: 0 15px; transition: transform 0.3s ease; cursor: pointer; border-radius: 5px; }}
    .social-logo:hover {{ transform: scale(1.3) translateY(-5px); }}
    
    /* Special Dark Container for White Beeclue Logo */
    .beeclue-logo-box {{
        background-color: #001c3d;
        padding: 20px;
        border-radius: 10px;
        display: inline-block;
        margin-top: 15px;
        width: 250px;
        transition: 0.3s;
    }}
    .beeclue-logo-box:hover {{ background-color: #003066; transform: scale(1.02); }}
    .beeclue-img {{ width: 100%; height: auto; }}
    
    .metric-title {{ color: #1a237e; font-weight: bold; margin-bottom: 5px; }}
</style>
""", unsafe_allow_html=True)

# ==========================================
# 3. HEADER (Main)
# ==========================================
st.markdown(f"""
<div class="header-box">
    <img src="{PSPCL_LOGO}" class="pspcl-main-logo">
    <h1 style="color: #1a237e; margin: 0; font-weight: 800;">PUNJAB STATE POWER CORPORATION LIMITED</h1>
    <h2 style="color: #444; margin: 5px; font-weight: 400;">11kV Voltage Drop Calculator Pro</h2>
</div>
""", unsafe_allow_html=True)

# ==========================================
# 4. SESSION STATE
# ==========================================
if 'final_df' not in st.session_state: st.session_state.final_df = None
if 'metrics' not in st.session_state: st.session_state.metrics = {}
if 'show_diagram' not in st.session_state: st.session_state.show_diagram = False

# ==========================================
# 5. SIDEBAR (Feeder Information)
# ==========================================
with st.sidebar:
    st.image(PSPCL_LOGO, width=120) # Side Logo
    st.title("⚙️ Control Panel")
    
    with st.expander("📌 Feeder Information", expanded=True):
        f_name = st.text_input("Feeder Name", value="")
        ss_name = st.text_input("Substation Name", value="")
        s_div = st.text_input("Sub-Division", value="")
        division = st.text_input("Division", value="")
    
    with st.expander("⚡ Parameters", expanded=True):
        mdi_amps = st.number_input("Max Demand (MDI Amps)", min_value=0.0, step=0.1, value=0.0)
        n_sec = st.number_input("Number of Sections", min_value=1, max_value=100, step=1, value=1)
        
        # Calculate MDI kVA
        mdi_kva = round(np.sqrt(3) * 11 * mdi_amps, 4)
        st.success(f"Calculated MDI: {mdi_kva} kVA")
    
    st.divider()
    st.caption(f"App Date: {datetime.now().strftime('%d-%m-%Y')}")

# ==========================================
# 6. DATA INPUT TABLE
# ==========================================
st.subheader("📍 Network Configuration")
# labels Source to Tail
sec_labels = [f"{chr(65+i)}-{chr(66+i)}" for i in range(n_sec)]
df_template = pd.DataFrame({
    "SECTION": sec_labels,
    "CONDUCTOR TYPE": [None] * n_sec,
    "LENGTH (KM)": [0.0] * n_sec,
    "NODE LOAD (kVA)": [0.0] * n_sec
})

user_input_df = st.data_editor(
    df_template,
    column_config={
        "SECTION": st.column_config.TextColumn("Span ID", disabled=True),
        "CONDUCTOR TYPE": st.column_config.SelectboxColumn("Conductor/Cable", options=list(VD_FACTORS.keys()), required=True),
        "LENGTH (KM)": st.column_config.NumberColumn("Length (km)", format="%.3f", min_value=0.0),
        "NODE LOAD (kVA)": st.column_config.NumberColumn("Node Load (kVA)", format="%.1f", min_value=0.0)
    },
    use_container_width=True, num_rows="fixed", key="vd_data_editor"
)

# ==========================================
# 7. CALCULATION CORE & BUTTONS
# ==========================================
c1, c2, c3 = st.columns(3)

def perform_calculations():
    with st.spinner("Processing network loads..."):
        time.sleep(0.3)
        calc_df = user_input_df.copy()
        
        # Cumulative Load (Tail to Source backward sum)
        node_loads = calc_df["NODE LOAD (kVA)"].tolist()
        cum_loads = [0] * len(node_loads)
        running_total = 0
        for i in range(len(node_loads)-1, -1, -1):
            running_total += node_loads[i]
            cum_loads[i] = running_total
            
        calc_df["CUMULATIVE LOAD (kVA)"] = cum_loads
        calc_df["VD FACTOR"] = calc_df["CONDUCTOR TYPE"].map(VD_FACTORS)
        calc_df["SECTION VD (V)"] = calc_df["LENGTH (KM)"] * calc_df["CUMULATIVE LOAD (kVA)"] * calc_df["VD FACTOR"]
        
        # Results Aggregation
        sum_vd = calc_df["SECTION VD (V)"].sum()
        max_ld_kva = cum_loads[0] if cum_loads else 0
        demand_f = mdi_kva / max_ld_kva if max_ld_kva > 0 else 0
        act_vd_volts = sum_vd * demand_f
        final_percent = (act_vd_volts / (11000 - act_vd_volts) * 100) if (11000 - act_vd_volts) > 0 else 0
        
        st.session_state.final_df = calc_df
        st.session_state.metrics = {
            "sum_vd": sum_vd, "max_ld_kva": max_ld_kva, "demand_f": demand_f,
            "act_vd_volts": act_vd_volts, "vd_percentage": final_percent, "mdi_kva": mdi_kva
        }
        st.session_state.show_diagram = False # Reset sketch view on calculate

if c1.button("🚀 EXECUTE CALCULATION", type="primary", use_container_width=True):
    if user_input_df["CONDUCTOR TYPE"].isnull().any():
        st.error("Error: Please select Conductor Type for all sections.")
    else:
        perform_calculations()

# ==========================================
# 8. RESULTS DISPLAY & METRICS
# ==========================================
if st.session_state.final_df is not None:
    res = st.session_state.final_df
    m = st.session_state.metrics
    
    st.subheader("📊 Final Calculation Table")
    st.dataframe(res, use_container_width=True)
    
    # Formulas Box
    st.markdown('<div class="formula-section">', unsafe_allow_html=True)
    m_col1, m_col2, m_col3 = st.columns(3)
    
    with m_col1:
        st.latex(r"D.F. = \frac{MDI_{kVA}}{\sum Load_{kVA}}")
        st.metric("Demand Factor", f"{m['demand_f']:.4f}")
    with m_col2:
        st.latex(r"Actual\ VD(V) = \sum VD \times DF")
        st.metric("Actual VD (Volts)", f"{m['act_vd_volts']:.2f} V")
    with m_col3:
        st.latex(r"\% VD = \frac{VD}{11000-VD} \times 100")
        st.metric("Final Drop (%)", f"{m['vd_percentage']:.3f} %")
    st.markdown('</div>', unsafe_allow_html=True)

    # Secondary Trigger: Show Diagram
    if c2.button("🎨 GENERATE SKETCH & REPORTS", use_container_width=True):
        st.session_state.show_diagram = True

    # EXPORT RAW CSV
    csv_raw = res.to_csv(index=False).encode('utf-8')
    c3.download_button("📊 EXPORT RAW DATA (CSV)", csv_raw, f"{feeder_name}_Raw.csv", "text/csv", use_container_width=True)

# ==========================================
# 9. SKETCH & OFFICIAL REPORTS (TRIGGERED)
# ==========================================
if st.session_state.show_diagram and st.session_state.final_df is not None:
    res_df = st.session_state.final_df
    m_res = st.session_state.metrics
    
    st.subheader("🎨 Feeder Single Line Diagram (Sketch)")
    sld = graphviz.Digraph()
    sld.attr(rankdir='BT') # Tail to Source view
    
    # Source Node
    source_lbl = f'{ss_name}\n(11kV Source)\n{m_res["max_ld_kva"]} kVA'
    sld.node('SOURCE', source_lbl, shape='house', style='filled', fillcolor='gold')
    
    prev_n = 'SOURCE'
    for idx, row in res_df.iterrows():
        curr_n = f"NODE_{idx}"
        is_c = "CABLE" in row['CONDUCTOR TYPE'].upper()
        # Shape: Box for Cable, Circle for ACSR
        sld.node(curr_n, row['SECTION'].split('-')[-1], shape='box' if is_c else 'circle', style='filled', fillcolor='#e1f5fe')
        edge_lbl = f"{row['CONDUCTOR TYPE']}\n{row['LENGTH (KM)']}km"
        sld.edge(curr_n, prev_n, label=edge_lbl, color='red' if is_c else 'black')
        prev_n = curr_n
    st.graphviz_chart(sld)

    # ==========================================
    # 10. ADVANCED EXPORT BUTTONS
    # ==========================================
    st.subheader("📥 Export Official Reports")
    e_col1, e_col2 = st.columns(2)
    
    # --- PDF EXPORT (fpdf2 style) ---
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 16)
    pdf.cell(0, 15, "PSPCL VOLTAGE DROP REPORT", ln=True, align='C')
    pdf.ln(5)
    
    pdf.set_font("helvetica", '', 11)
    pdf.cell(95, 8, f"Feeder: {feeder_name}")
    pdf.cell(95, 8, f"Sub-Division: {s_div}", ln=True)
    pdf.cell(95, 8, f"Substation: {ss_name}")
    pdf.cell(95, 8, f"Division: {division}", ln=True)
    pdf.cell(95, 8, f"MDI: {m_res['mdi_kva']} kVA", ln=True)
    pdf.ln(5)
    
    # Table Header
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font("helvetica", 'B', 10)
    for h, w in [("Section", 30), ("Conductor", 55), ("Len", 25), ("Cum.kVA", 35), ("Sec VD", 45)]:
        pdf.cell(w, 10, h, 1, 0, 'C', True)
    pdf.ln()
    
    # Table Body
    pdf.set_font("helvetica", '', 9)
    for _, r in res_df.iterrows():
        pdf.cell(30, 8, str(r['SECTION']), 1)
        pdf.cell(55, 8, str(r['CONDUCTOR TYPE'])[:20], 1)
        pdf.cell(25, 8, str(r['LENGTH (KM)']), 1)
        pdf.cell(35, 8, f"{r['CUMULATIVE LOAD (kVA)']:.1f}", 1)
        pdf.cell(45, 8, f"{r['SECTION VD (V)']:.3f}", 1, 1)
        
    pdf.ln(10)
    pdf.set_font("helvetica", 'B', 14)
    pdf.set_text_color(200, 0, 0)
    pdf.cell(0, 10, f"FINAL PERCENTAGE VOLTAGE DROP: {m_res['vd_percentage']:.3f} %", ln=True)
    
    # signature
    pdf.ln(30)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(125)
    pdf.multi_cell(60, 6, f"__________________\nSDO OP SUB-DIVISION\nPSPCL {s_div.upper()}", align='C')
    
    pdf_out = pdf.output() # fpdf2 output
    e_col1.download_button("📥 DOWNLOAD OFFICIAL PDF", bytes(pdf_out), f"{feeder_name}_Report.pdf", "application/pdf", use_container_width=True)

    # --- Excel Export (Integrated Your Code) ---
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary Report"
    # Basic Summary sheet
    ws_summary["A1"] = f"Feeder: {feeder_name}"
    ws_summary["A2"] = f"Total Load: {m_res['max_ld_kva']} kVA"
    ws_summary["A3"] = f"Actual VD: {m_res['act_vd_volts']} V"
    ws_summary["A4"] = f"Regulation: {m_res['vd_percentage']} %"
    
    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    e_col2.download_button("📥 DOWNLOAD EXCEL REPORT", excel_buf.getvalue(), f"{feeder_name}_Report.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

# ==========================================
# 11. FOOTER SECTION (Branding Fully Fixed)
# ==========================================
st.markdown(f"""
<div class="footer-container">
    <p style="font-size: 1.4em; font-weight: bold;">Made with ❤️ by Anuj Narang</p>
    <p style="color: #666; margin-bottom: 30px;">Junior Engineer (Electrical) | PSPCL Professional Development</p>
    
    <div style="margin-bottom: 40px;">
        <a href="https://instagram.com/iamanujnarang" target="_blank"><img src="{INSTA_ICON}" class="social-logo" alt="Insta"></a>
        <a href="https://facebook.com/iamanujnarang" target="_blank"><img src="{FB_ICON}" class="social-logo" alt="FB"></a>
        <a href="https://x.com/iamanujnarang" target="_blank"><img src="{X_LOGO_FINAL}" class="social-logo" alt="X"></a>
        <a href="https://linkedin.com/in/iamanujnarang" target="_blank"><img src="{LINKEDIN_ICON}" class="social-logo" alt="LinkedIn"></a>
    </div>
    
    <p style="color: #999; font-size: 0.9em; text-transform: uppercase;">In Strategic Collaboration with</p>
    
    <div class="beeclue-logo-box">
        <a href="https://beeclue.com" target="_blank">
            <img src="{BEECLUE_LOGO_WHITE}" class="beeclue-img" alt="Beeclue">
        </a>
    </div>
    
    <p style="color: #bbb; font-size: 0.8em; margin-top: 30px;">© 2026 Anuj Narang. All Rights Reserved.</p>
</div>
""", unsafe_allow_html=True)
